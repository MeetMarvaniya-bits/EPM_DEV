import calendar
from salary_manage import Salarymanage
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.styles import numbers


class SalaryData():

    def __init__(self, db):
        self.db = db

    def add_data(self, salid, companyname):

        # # FOR CUSTOM DETAIL 1
        mont_in_num = int(salid.split('_')[0][5:])
        month = calendar.month_name[mont_in_num]
        year = int(salid.split('_')[1])

        # # FOR VALUE DATE
        next_month = 0
        next_year = 0

        if mont_in_num == 12:
            next_month += 1
            next_year += year + 1
        else:
            next_month += mont_in_num + 1
            next_year += year

        value_date = f'01/{next_month}/{next_year}'

        debited_account_no = '02940200001115'

        custom_detail_1 = f'{month}{year}'

        custom_detail_2 = ''

        custom_detail_3 = ''

        custom_detail_4 = ''

        custom_detail_5 = ''

        custom_detail_6 = ''

        remark = 'ALIAN SOFTWARE SALARY'

        purpose = 'SALARY'

        wb = Workbook()

        ws = wb.active
        ws.title = "Sheet"

        # # Set Header
        title_row = ['CUSTOM_DETAILS1', 'Value Date', 'Message Type', 'Debit Account No.', 'Beneficiary Name',
                     'Payment Amount', 'Beneficiary Bank Swift Code / IFSC Code', 'Beneficiary Account No.',
                     'Transaction Type Code', 'CUSTOM_DETAILS2', 'CUSTOM_DETAILS3', 'CUSTOM_DETAILS4',
                     'CUSTOM_DETAILS5', 'CUSTOM_DETAILS6', 'Remarks', 'Purpose of Payment']

        ws.append(title_row)

        # # SET CELL WIDTH
        for n in range(1, 17):
            ws.column_dimensions[ws.cell(row=1, column=n).column_letter].width = 20

        # # GET ALL EMPLOYEE DATA
        salary_list = Salarymanage(self.db).get_all_emp_salary_data(companyname, salid)

        for i in salary_list:

            empid = salary_list[i]["userID"]

            user_ref = self.db.collection(companyname).document('employee').collection('employee').document(empid)

            data = user_ref.get().to_dict()

            benificiary_bank = data['bankName']
            if benificiary_bank == "BOB":

                message_type = 'IFT'

                transaction_type = 'IFT'

            else:

                message_type = 'NEFT'

                transaction_type = 'NEFT'

            benificiary_name = data['employeeName'].capitalize()
            net_salary = salary_list[i]['netSalary']

            if net_salary == '':

                payment_amount = '0.00'

            else:

                payment_amount = f'{round(float(net_salary), 2)}'
            if data['bankName'] != "BOB":
                # beneficiary_ifsc_code = data['ifscCode']
                beneficiary_ifsc_code = "skdjbvskjdb"
            else:
                beneficiary_ifsc_code = ""

            beneficiary_account_no = data['accountNumber']
            # beneficiary_account_no = "9874987987987"

            data = {'CUSTOM_DETAILS1': custom_detail_1, 'Value Date': value_date, 'Message Type': message_type,
                    'Debit Account No.': debited_account_no, 'Beneficiary Name': benificiary_name,
                    'Payment Amount': payment_amount, 'Beneficiary Bank Swift Code / IFSC Code': beneficiary_ifsc_code,
                    'Beneficiary Account No.': beneficiary_account_no, 'Transaction Type Code': transaction_type,
                    'CUSTOM_DETAILS2': custom_detail_2, 'CUSTOM_DETAILS3': custom_detail_3,
                    'CUSTOM_DETAILS4': custom_detail_4, 'CUSTOM_DETAILS5': custom_detail_5,
                    'CUSTOM_DETAILS6': custom_detail_6, 'Remarks': remark, 'Purpose of Payment': purpose
                    }

            employee_data = [data[title_row[0]], data[title_row[1]], data[title_row[2]], data[title_row[3]],
                             data[title_row[4]], data[title_row[5]], data[title_row[6]], data[title_row[7]],
                             data[title_row[8]], data[title_row[9]], data[title_row[10]], data[title_row[11]],
                             data[title_row[12]], data[title_row[13]], data[title_row[14]], data[title_row[15]]]

            ws.append(employee_data)

        # # SET FORMATE AS A TEXT
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = numbers.FORMAT_TEXT

        # # TEXT ALIGN LEFT
        alignment = Alignment(horizontal='left')
        for n in range(1, len(salary_list) + 1):
            for cell in ws[str(n)]:
                cell.alignment = alignment
        print("done")
        return wb
