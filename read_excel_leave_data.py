import time
from openpyxl import load_workbook
from datetime import datetime, timedelta


def read_excel_rows(file):
    workbook = load_workbook(filename=file)

    print('workbook')

    sheet = workbook['Sheet']

    max_row = sheet.max_row

    data = []

    widgets = [
        'Processing: ',
        progressbar.Percentage(),
        ' ',
        progressbar.Bar(marker='█'),
        ' ',
        progressbar.ETA()
    ]

    progress_bar = progressbar.ProgressBar(maxval=max_row, widgets=widgets).start()

    for row in range(1, max_row + 1):
        row_data = []
        for column in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=column).value
            # Check if the cell value is a datetime object
            if isinstance(cell_value, datetime):
                # Convert the datetime to the desired format
                cell_value = cell_value.strftime("%d-%m-%Y")

            if cell_value != None:
                row_data.append(cell_value)

        data.append(row_data)
    progress_bar.finish()

    header_row_index = None
    attendance_start_index = None

    for i, row in enumerate(data):
        if row == ['Sr No', 'User ID', 'Name ', 'PR', 'WO', 'PH', 'PL', 'TR', 'AB', 'UL', 'FB', 'RD', 'LI', 'EO', 'OT',
                   'Auth OT', 'Auth C-OFF', 'WrkHrs']:
            header_row_index = i
        elif row == ['Shift', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS',
                     'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS', 'GS',
                     'GS']:
            attendance_start_index = i - 1

    #
    # if header_row_index is None or attendance_start_index is None:
    #     print("Invalid data format. Unable to find header row or attendance data.")
    #     exit()

    column_headers_personal = data[header_row_index]

    column_headers_for_hours = data[attendance_start_index]

    personal_records = []
    attendance_hours = []
    attendance_record_1 = []
    attendance_record_2 = []
    for row in data:
        if row != []:

            if isinstance(row[0], int):
                record = dict(zip(column_headers_personal, row))
                personal_records.append(record)

            # # FOR GET HOURS DATE WISE
            # elif row[0].startswith('WrkHrs'):
            #     record = dict(zip(column_headers_for_hours, row[1:]))
            #     attendance_hours.append(dict(record))

            elif row[0].startswith('WrkHrs'):
                record = row[1:]
                # print(record)
                total_time = timedelta()
                for t in record:
                    if t != '-':
                        time_parts = t.split(':')
                        hours = int(time_parts[0])
                        minutes = int(time_parts[1])
                        time_delta = timedelta(hours=hours, minutes=minutes)
                        total_time += time_delta
                total_hours = int(total_time.total_seconds() // 3600)
                total_minutes = int((total_time.total_seconds() % 3600) // 60)
                total_time = {'WrkHrs': f'{total_hours}:{total_minutes}'}
                attendance_hours.append(total_time)

            elif row[0].startswith('Stat1'):
                record = dict(zip(column_headers_for_hours, row[1:]))
                attendance_record_1.append(dict(record))

            elif row[0].startswith('Stat2'):
                record = dict(zip(column_headers_for_hours, row[1:]))
                attendance_record_2.append(dict(record))

    attendance_records = []

    for d1, d2 in zip(attendance_record_1, attendance_record_2):
        PL = 0
        CL = 0
        SL = 0
        new_dict_1 = d1.copy()
        for key, value in new_dict_1.items():
            if value == 'CL':
                CL += 1
            elif value == 'PL':
                PL += 1
            elif value == 'SL':
                SL += 1
        new_dict_2 = d2.copy()
        for key, value in new_dict_2.items():
            if value == 'CL':
                CL += 1
            elif value == 'PL':
                PL += 1
            elif value == 'SL':
                SL += 1
        total_leave = {'C_L': CL / 2, 'P_L': PL / 2, 'S_L': SL / 2}
        attendance_records.append(total_leave)

    new_data = []

    for d1, d2, d3 in zip(personal_records, attendance_hours, attendance_records):
        new_dict = d1.copy()
        new_dict.update(d2)
        new_dict.update(d3)
        new_data.append(new_dict)

    return new_data
