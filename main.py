import datetime
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, render_template, request, redirect, url_for, session, send_file, Response
import io
from leave_manage import Leavemanage
from firebase_admin import credentials,  auth, exceptions, firestore
from details import Profile
from create_new_employee import Create
from salary_manage import Salarymanage
from department import Department
from update_employee import Update_information
from dashboard import Dashboard
import re
from excel_sheet import SalaryData
import concurrent.futures
from salary_slip import SalarySlip
from register import Register
from admin_register import Admin_Register
from login import Login
from moth_days import MonthCount
from mail import Mail
import concurrent.futures
from salary_calculation import SalaryCalculation
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from generate_excel import create_excel_file
from read_data import ExcelData
from apscheduler.schedulers.background import BackgroundScheduler
import os
import platform
import read_excel_leave_data
import json
import requests
from functools import wraps
import os
import platform



# FLASK APP
app = Flask(__name__)
app.secret_key = 'tO$&!|0wkamvVia0?n$NqIRVWOG'
app.config['SESSION_TYPE'] = 'filesystem'
app.permanent_session_lifetime = datetime.timedelta(hours=24)
# USE A SERVICE ACCOUNT
cred = credentials.Certificate('empoyee-payroll-system-firebase-adminsdk-5h89d-1602329ca8.json')
db = firestore.client()
leaveobj = Leavemanage(db)
dept = Department(db)
update_obj = Update_information(db)
dashboard_obj = Dashboard(db)
register_obj = Register(db)
admin_register_obj = Admin_Register(db)
login_obj = Login(db)
moth_count = MonthCount()
mail_obj = Mail()
companyname='alian_software'
FIREBASE_WEB_API_KEY = "AIzaSyDe2qwkIds8JwMdLBbY3Uw7JQkFRNXtFqo"
rest_api_url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword"


SECOND_FIREBASE_API_KEY = "AIzaSyB_RwtduOw9glo9UIoL-S7ng7b2LKY7iDo"
# Testing path
# C:/Users/alian/Desktop/Testing
# C:/Users/alian/Downloads/my_file.xlsx


def clear_session_data():
    with app.app_context():
        session.clear()

scheduler = BackgroundScheduler()
scheduler.add_job(clear_session_data, 'interval', days=1, start_date=datetime.datetime.now().replace(hour=0, minute=0, second=0))
scheduler.start()


def sign_in_with_email_and_password(email: str, password: str, return_secure_token: bool = True):
    payload = json.dumps({
        "email": email,
        "password": password,
        "returnSecureToken": return_secure_token
    })

    r = requests.post(rest_api_url,
                      params={"key": FIREBASE_WEB_API_KEY},
                      data=payload)

    return r.json()



def SECOND_sign_in_with_email_and_password(email: str, password: str, return_secure_token: bool = True):
    payload = json.dumps({
        "email": email,
        "password": password,
        "returnSecureToken": return_secure_token
    })

    r = requests.post(rest_api_url,
                      params={"key": SECOND_FIREBASE_API_KEY},
                      data=payload)

    return r.json()



def login_required(route_function):
    @wraps(route_function)
    def wrapper(*args, **kwargs):
        if 'auth_user_id' in session:
            # User is logged in, proceed with the route function
            return route_function(*args, **kwargs)
        else:
            # User is not logged in, redirect to login page
            return redirect(url_for('login'))
    return wrapper


@app.route('/', methods=["POST", "GET"])
def login():
    global companyname
    responce = ''
    if request.method == 'POST':
        data = request.form

        result = {}
        for key, value in data.items():
            result[key] = value
        print(result)

        user_auth = sign_in_with_email_and_password(email=result['email'], password=result['password'])
        if "registered" in user_auth:
            # print(user_auth['registered'])
            # print(user_auth['localId'])
            print(user_auth)
            session['auth_user_id'] = user_auth['localId']
            responce = login_obj.login(user_auth, companyname)
        else:
            print(user_auth["error"]["message"])



        if responce == 'Admin':
            return redirect(url_for('employee_list', username=responce))
        elif responce == 'HR':
            return redirect(url_for('employee_list', username=responce))

        elif responce == "Employee":
            return redirect(url_for('employee_list', username=responce))

        else:
            responce = 'Inavalid Id and Password'
    company_list = []
    collections = db.collections()
    for collection in collections:
        company_list.append(collection.id)
    ''' LOGIN PAGE '''
    url = f'/'
    return render_template('login.html', responce=responce, url=url, company_list=company_list)

@app.route('/logout')
def logout():
    # Clear all session variables
    session.clear()

    # Redirect to the login page
    return redirect(url_for('login'))


@app.route('/forgot_password', methods=["POST", "GET"])
def forgot_password():
    if request.method == 'POST':
        print(request.form.to_dict())
        email = request.form.to_dict()['email']
        if email != None:
            Admin = db.collection(companyname).document('admin').get().to_dict()
            if Admin != None:
                company_mail = Admin['AdminID']
                auth_password = Admin['auth_password']
                if company_mail == email:
                    email = email
                    password = Admin['password']
                    mail_obj.forgot_mail(email=email, password=password,
                                         company_mail=company_mail, auth_password=auth_password)
                else:
                    user = db.collection(companyname).document('employee').collection('employee').where(
                        'email', "==", email).get()
                    if user != None:
                        data = user[0].to_dict()
                        email = email
                        password = data['password']
                        mail_obj.forgot_mail(email=email, password=password,
                                             company_mail=company_mail, auth_password=auth_password)

    return redirect(url_for('login'))

@app.route('/success', methods=["POST", "GET"])
def success():
    return render_template('success.html')


# @app.route('/', methods=['GET', 'POST'])
# def register():
#     responce = ''
#     if request.method == 'POST':
#         data = request.form
#         companyname = data["CompanyName"]
#         responce = register_obj.register(data)
#         companyname=companyname
#
#         if responce == True:
#             return redirect(url_for('login', ))
#
#     ''' REGISTER PAGE '''
#     return render_template('register.html', responce=responce)



@app.route('/admin_register', methods=['GET', 'POST'])
def admin_register():
    global companyname
    responce = ''
    if request.method == 'POST':
        data = request.form
        responce = admin_register_obj.admin_register(data, companyname)

        if responce == True:

            return redirect(url_for('login', ))

    ''' REGISTER PAGE '''
    return render_template('admin_register.html', responce=responce)

@app.route('/<username>/dashboard', methods=['GET', 'POST'])
@login_required
def dashboard(username):
    session.pop('excel_path', default=None)

    if request.method == 'POST':
        form = request.form.to_dict()

        data = {str(form["date"]): form["description"]}
        db.collection(companyname).document('holidays').update(data)
        holidays = db.collection(companyname).document('holidays').get().to_dict()
        moath_data = moth_count.count(holidays)
        db.collection(companyname).document('month_data').set(moath_data)

    holidays = db.collection(companyname).document('holidays').get().to_dict()
    moath_data = moth_count.count(holidays)
    working_days = moath_data['workingDays']
    # Check the current date
    if datetime.datetime.now().day == 1 and 'session_leave' not in session:
        leaveobj.leave_add(companyname)
        session['session_leave'] = datetime.datetime.now()
        print("create session Variable")


    if datetime.datetime.today().day == 1 and datetime.datetime.today().month == 1:

        leaveobj.leave_reset(companyname)

    ''' DISPLAY DASHBOARD '''
    dashboard_data = dashboard_obj.all_data(companyname)
    holidays = moth_count.get_holidays(holidays)

    employee_data = []
    with concurrent.futures.ThreadPoolExecutor() as executor:
        users_ref = db.collection(companyname).document('employee').collection('employee')
        for emp_doc in users_ref.stream():
            employee_data.append(executor.submit(dashboard_obj._get_employee_data, emp_doc))
    for future in employee_data:
        print(future.result())
    employee_on_leave, total_leaves, employee_birthday, employee_anniversary = {}, {}, {}, {}
    for future in concurrent.futures.as_completed(employee_data):
        result = future.result()
        if result is not None:

            if 'birthday' in result:
                employee_birthday[result['name']] = result['birthday']

                # employee_birthday = sorted(employee_birthday.items(), key=lambda x: x[1], reverse=True)
            if 'anniversary' in result:
                employee_anniversary[result['name']] = result['anniversary']

            if result['leaves']:
                employee_on_leave[result['name']] = result['leaves']

            if result['total_leaves'] > 0:
                total_leaves[result['name']] = result['total_leaves']

    total_leaves = sorted(total_leaves.items(), key=lambda x: x[1], reverse=True)
    total_leaves = dict(total_leaves[:5])
    print(employee_on_leave)
    # Get current month and year

    # Print the resulting dictionary

    return render_template('dashboard.html', employee_on_leave=employee_on_leave, total_leaves=total_leaves,
                           employee_birthday=employee_birthday, employee_anniversary=employee_anniversary,
                           moath_data=moath_data, holidays=holidays, username=username, dashboard_data=dashboard_data)


@app.route('/<username>/employeelist', methods=['GET', 'POST'])
@login_required
def employee_list(username):
    if 'increment' not in session:
        increments=db.collection('alian_software').document('increments').get().to_dict()['increments']
        for increment in increments:
            converted_date = datetime.datetime.strptime(increment['effectiveDate'], '%Y-%m-%d').date()
            if datetime.datetime.today().date()==  converted_date:
                print(increment['empid'],(increment['total']))
                db.collection('alian_software').document('employee').collection('employee').document(increment['empid']).update({'salary':round(float(increment['total']))* 12})
        session['increment'] = 'Done'
    # SENDING EMPLOYEE MAIL FOR ADD DETAILS
    if request.method == 'POST':
        employee_mail = request.form.get('new_email')
        auth_data = db.collection(companyname).document('admin').get().to_dict()
        company_mail = auth_data['AdminID']
        auth_password = auth_data['auth_password']
        mail_obj.new_employee_mail(email=employee_mail, company_mail=company_mail, auth_password=auth_password)

    ''' DISPLAY LIST OF EMPLOYEES IN COMPANY '''

    # CALCULATE EXPERIENCE OF EMPLOYEE
    def calculate_experience(data):
        new_data = data
        try:
            doj = (data['doj'])
            today_date = datetime.datetime.today().date()
            start_date = datetime.datetime.strptime(doj, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(str(today_date), '%Y-%m-%d')
            if start_date > end_date:
                # Extract the years, months, and days from the delta
                years = 0
                months = 0
                days = 0
                current_experience = {'currentExperience': f'{years} Year(s) {months} Month(s) {days} Day(s)'}
                db.collection(companyname).document(u'employee').collection('employee').document(data['userID']).update(current_experience)
            else:
                # Calculate the difference
                delta = end_date - start_date
                # Extract the years, months, and days from the delta
                years = delta.days // 365
                months = (delta.days % 365) // 30
                days = (delta.days % 365) % 30
                current_experience = {'currentExperience': f'{years} Year(s) {months} Month(s) {days} Day(s)'}
                db.collection(companyname).document(u'employee').collection('employee').document(data['userID']).update(current_experience)
            new_data.update(current_experience)
        except:
            pass
        return new_data

    # GET ALL EMPLOYEE DATA
    def get_employee_data():

        docs = db.collection(companyname).document(u'employee').collection('employee').stream()
        employee_list = {}
        for doc in docs:
            if 'user_status' not in doc.to_dict():
                if doc.to_dict()['role'] == 'Admin':
                    continue
                employee_list.update({doc.id: doc.to_dict()})
            elif doc.to_dict()['user_status'] != 'disable':
                if doc.to_dict()['role'] == 'Admin':
                    continue
                employee_list.update({doc.id: doc.to_dict()})
        return employee_list


    # GET DEPARTMENTS IN ORGANIZAT  ION
    def get_department_data():
        department = (db.collection(companyname).document(u'department').get()).to_dict()
        return department

    with concurrent.futures.ThreadPoolExecutor() as executor:
        employee_data = executor.submit(get_employee_data)
        department_data = executor.submit(get_department_data)
    employee_list = employee_data.result()
    department = department_data.result()
    print(employee_list)
    return render_template('employees_list.html', data=employee_list, department=department, username=username)


@app.route("/<username>/storage-path", methods=["POST"])
@login_required
def excel_sheet_path(username):
    excel_path = request.json["excel_path"]
    excel = ExcelData(db)
    excel.store_excel_data(companyname, excel_path)
    return redirect(url_for('dashboard', username=username))


@app.route('/<username>/result', methods=['POST', 'GET'])
@login_required
def add(username):

    ''' NEW EMPLOYEE DATA STORE IN DATABASE AND DISPLAY IN LIST '''
    create = Create(db, companyname)
    # create.result()
    employee_mail = request.form.get('email')
    employee_mail = request.form.get('email')
    password = request.form.get('password')
    new_id = create.result()
    auth_data = db.collection(companyname).document('admin').get().to_dict()
    company_mail = auth_data['AdminID']
    auth_password = auth_data['auth_password']
    mail_obj.employee_registered_mail(email=employee_mail, password=password, company_mail=company_mail, auth_password=auth_password)

    return redirect(url_for('employee_list',username=username))



@app.route('/<username>/<id>/delete', methods=['POST', 'GET'])
@login_required
def delete_employee(username, id):
    doc_ref = db.collection(companyname).document(u'employee').collection('employee').document(id)
    doc_ref.update({'user_status': 'disable'})
    return redirect(url_for('employee_list', username=username))


@app.route('/register_employee', methods=['POST', 'GET'])
@login_required
def employee_register_by_mail():

    # GET ALL EMPLOYEES DETAILS
    def get_employee_data():

        docs = db.collection(companyname).document(u'employee').collection('employee').stream()
        employee_list = {}
        for doc in docs:
            if 'user_status' not in doc.to_dict():
                employee_list.update({doc.id: doc.to_dict()})
            elif doc.to_dict()['user_status'] != 'disable':
                employee_list.update({doc.id: doc.to_dict()})
        return employee_list

    data = get_employee_data()

    responce = ''
    if request.method == 'POST':
        data = request.form
        email = data["email"]

        doc = db.collection(companyname).document("employee").collection('employee').where('email', "==", email).get()
        if len(doc) > 0:
            responce = 'Already Email Exist Try Different Email'
        else:

            ''' NEW EMPLOYEE DATA STORE IN DATABASE AND DISPLAY IN LIST '''
            create = Create(db, companyname)
            create.result()
            employee_mail = request.form.get('email')
            auth_data = db.collection(companyname).document('admin').get().to_dict()
            company_mail = auth_data['AdminID']
            auth_password = auth_data['auth_password']
            mail_obj.employee_registered_mail(employee_mail, companyname, company_mail, auth_password)

            return redirect(url_for('login'))

    def get_department_data():
        department = (db.collection(companyname).document(u'department').get()).to_dict()
        return department

    with concurrent.futures.ThreadPoolExecutor() as executor:
        department_data = executor.submit(get_department_data)
    department = get_department_data()
    return render_template('add_employee.html',department=department,
                           department_data=department_data, responce=responce, data=data)

@app.route('/create_excel')
@login_required
def create_excel():
    ''' EXCEL SHEET DATA FORMATE FOR NEW COMPANY '''
    # Create the Excel file
    wb = create_excel_file()

    # Save the file to a BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return the file as a response with appropriate headers
    return Response(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": "attachment;filename=my_file.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    )


@app.route('/upload_file', methods=['POST'])
@login_required
def upload_file(companyname):
    if request.method == 'POST':
        file = request.files['file']
        if file.filename != '':
            workbook = load_workbook(filename=file)
            worksheet = workbook.active
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(row)
            return redirect(url_for('employee_list', ))
    return redirect(url_for('employee_list', ))


@app.route('/<username>/employeeprofile/<id>', methods=['GET', 'POST'])
@login_required
def employee_profile(username, id):
    ''' DISPLAY EMPLOYEE DETAILS '''
    users_ref = db.collection(str(companyname)).document('employee').collection('employee').document(id).collection(
        'leaveMST')
    if request.method == 'POST':
        ''' Store leave Data '''
        result = request.form

        with ThreadPoolExecutor(max_workers=2) as executor:
            executor.submit(leaveobj.take_leave, users_ref, data=result)

    ''' GET LEAVE DATA '''
    start=datetime.datetime.now()
    with ThreadPoolExecutor(max_workers=4) as executor:
        total_leave_future = executor.submit(leaveobj.get_total_leave, users_ref)
        leave_list_future = executor.submit(leaveobj.leave_list, users_ref)

    total_leave = total_leave_future.result()
    leave_list = leave_list_future.result()
    end=datetime.datetime.now()
    print(end-start)
    ''' GET EMPLOYEE DATA '''
    start=datetime.datetime.now()

    with ThreadPoolExecutor(max_workers=4) as executor:
        personal_data_future = executor.submit(Profile(db, id, companyname).personal_data)
        tds_data_future = executor.submit(Profile(db, id, companyname).tds_data)
        salary_data_future = executor.submit(Profile(db, id, companyname).salary_data)

    increment_data = []
    personal_data = personal_data_future.result()
    end = datetime.datetime.now()
    print(end - start)

    start=datetime.datetime.now()
    for key,value in personal_data.items():
        if key.startswith('increment'):
            increment_data.append({key: value})


    increment_data = sorted(increment_data, key=lambda x: x[list(x.keys())[0]]['effectiveDate'])
    increment_list = [(index, content) for index, content in enumerate(increment_data)]
    print(increment_list)

    keys = [next(iter(item[1])) for item in increment_list]
    print(keys)
    effective_dates = [item[1][keys[i]]['effectiveDate'] for i, item in enumerate(increment_list)]
    print(effective_dates)

    contract_data = []
    personal_data = personal_data_future.result()
    for key,value in personal_data.items():
        if key.startswith('contract'):
            contract_data.append({key: value})
    contract_list = [(index, content) for index, content in enumerate(contract_data)]

    data = {'personal_data': personal_data_future.result(), 'tds_data': tds_data_future.result(),
            'salary_data': salary_data_future.result()}

    end = datetime.datetime.now()
    print(end - start)
    start=datetime.datetime.now()

    leave_status = False
    leave_status_date = ''
    for i in leave_list:
        if i != 'total_leaves':
            leave_date = (datetime.datetime.strptime(leave_list[i]['applydate'], '%Y-%m-%d').month)
            if leave_date == datetime.datetime.today().month:
                leave_status = True
                leave_status_date = (f'{leave_list[i]["fromdate"]} to {leave_list[i]["todate"]} ')

    end = datetime.datetime.now()
    print(end - start)

    def get_department_data():
        department = (db.collection(companyname).document(u'department').get()).to_dict()
        return department

    with concurrent.futures.ThreadPoolExecutor() as executor:
        department_data = executor.submit(get_department_data)
    department = department_data.result()


    return render_template('employee_profile.html', leave=leave_status, data=data, total_leave=total_leave,
                           leave_list=leave_list, leave_date=leave_status_date,username=username, department=department,
                           increment_data=increment_list, contract_data=contract_list, effective_dates=effective_dates)





@app.route('/<username>/employee_view/<id>', methods=['GET', 'POST'])
@login_required
def employee_view(username, id):
    ''' DISPLAY EMPLOYEE DETAILS '''
    users_ref = db.collection(str(companyname)).document('employee').collection('employee').document(id).collection(
        'leaveMST')
    if request.method == 'POST':
        ''' Store leave Data '''
        result = request.form

        with ThreadPoolExecutor(max_workers=2) as executor:
            executor.submit(leaveobj.take_leave, users_ref, data=result)

    ''' GET LEAVE DATA '''
    with ThreadPoolExecutor(max_workers=2) as executor:
        total_leave_future = executor.submit(leaveobj.get_total_leave, users_ref)
        leave_list_future = executor.submit(leaveobj.leave_list, users_ref)

    total_leave = total_leave_future.result()
    leave_list = leave_list_future.result()

    ''' GET EMPLOYEE DATA '''
    with ThreadPoolExecutor(max_workers=2) as executor:
        personal_data_future = executor.submit(Profile(db, id, companyname).personal_data)
        tds_data_future = executor.submit(Profile(db, id, companyname).tds_data)
        salary_data_future = executor.submit(Profile(db, id, companyname).salary_data)

    increment_data = []
    personal_data = personal_data_future.result()
    for key,value in personal_data.items():
        if key.startswith('increment'):
            increment_data.append({key: value})
    increment_list = [(index, content) for index, content in enumerate(increment_data)]

    contract_data = []
    personal_data = personal_data_future.result()
    for key,value in personal_data.items():
        if key.startswith('contract'):
            contract_data.append({key: value})
    contract_list = [(index, content) for index, content in enumerate(contract_data)]

    data = {'personal_data': personal_data_future.result(), 'tds_data': tds_data_future.result(),
            'salary_data': salary_data_future.result()}

    leave_status = False
    leave_status_date = ''
    for i in leave_list:
        if i != 'total_leaves':
            leave_date = (datetime.datetime.strptime(leave_list[i]['applydate'], '%Y-%m-%d').month)
            if leave_date == datetime.datetime.today().month:
                leave_status = True
                leave_status_date = (f'{leave_list[i]["fromdate"]} to {leave_list[i]["todate"]} ')

    return render_template('employee_view.html', leave=leave_status, data=data, total_leave=total_leave,
                           leave_list=leave_list, leave_date=leave_status_date,username=username,
                           increment_data=increment_list, contract_data=contract_list)


@app.route('/<username>/pdf/<id>/<salid>')
@login_required
def pdf_personal(username, id, salid):
    ''' SALARY SLIP PDF GENERATION '''
    path = get_download_folder()
    salary = SalarySlip(db)
    responce=salary.salary_slip_personal(companyname, id, salid, path)
    # CHECK THE USER
    return responce


@app.route('/<username>/personal_data_update/<id>', methods=['GET', 'POST'])
@login_required
def personal_data_update(username, id):
    """ UPDATE EMPLOYEE PERSONAL DETAILS """
    if request.method == 'POST':
        form = request.get_json()
        update_obj.update_personal_info(companyname, form, id)
    return redirect(url_for('employee_profile', id=id, username=username))


@app.route('/save_data/<empid>/<username>', methods=['POST'])
@login_required
def save_data(empid, username):
    """ Update/Add Data in Contract And Increment """
    ''' GET EMPLOYEE DATA '''
    with ThreadPoolExecutor(max_workers=2) as executor:
        personal_data_future = executor.submit(Profile(db, empid, companyname).personal_data)

    # # INCREMENT DATA
    increment_data = []
    contract_data = []
    previous_salary=0
    personal_data = personal_data_future.result()
    for key, value in personal_data.items():
        if key.startswith('increment'):
            increment_data.append({key: value})
        elif key.startswith('contract'):
            contract_data.append({key: value})

    increment_data= sorted(increment_data, key=lambda x: x[1][list(x[1].keys())[0]]['effectiveDate'])
    increment_list = [(index, content) for index, content in enumerate(increment_data)]
    increment_keys = [next(iter(item[1])) for item in increment_list]
    increment_keys.sort()
    contract_list = [(index, content) for index, content in enumerate(contract_data)]
    contract_keys = [next(iter(item[1])) for item in contract_list]
    contract_keys.sort()

    if request.method == 'POST':
        data = request.form.to_dict()
        contract_increment_data = {}
        for n in range(0, max(len(increment_keys),len(contract_keys))+1):
            for key, value in data.items():
                if len(increment_keys)>n:
                    if key.startswith(increment_keys[n]):
                        new_data = { increment_keys[n]:
                                {
                                    f'grossSalary': data[f'{increment_keys[n]}_grossSalary'],
                                    f'jobPosition': data[f'{increment_keys[n]}_jobPosition'],
                                    f'effectiveDate': data[f'{increment_keys[n]}_effectiveDate'],
                                    f'increment': data[f'{increment_keys[n]}_increment'],
                                    f'incrementType': data[f'{increment_keys[n]}_increment_type'],
                                    f'total': data[f'{increment_keys[n]}_total'],
                                    f'note': data[f'{increment_keys[n]}_note']}
                            }
                        contract_increment_data.update(new_data)
                elif key.startswith(f'new_inc_'):

                    if data[f'new_inc_grossSalary'] != "" and data[f'new_inc_jobPosition'] != "":
                        key=0
                        if len(increment_keys)>0:
                            key=int(increment_keys[-1][-2:])
                        if key<9:
                            key=f'0{key+1}'
                        else:
                            key=key+1

                        new_data = {f'increment_{key}':

                            {
                                f'grossSalary': data[f'new_inc_grossSalary'],
                                f'jobPosition': data[f'new_inc_jobPosition'],
                                f'effectiveDate': data[f'new_inc_effectiveDate'],
                                f'increment': data[f'new_inc_increment'],
                                f'incrementType': data[f'new_drop_increment'],
                                f'total': data[f'new_inc_total'],
                                f'note': data[f'new_inc_note']}
                        }
                        contract_increment_data.update(new_data)
                elif len(contract_keys)>n:
                    if key.startswith(contract_keys[n]):
                        new_data = {contract_keys[n]:
                            {
                                f'contractDate': data[f'{contract_keys[n]}_contractDate'],
                                f'contractPeriod': data[f'{contract_keys[n]}_contractPeriod'],
                                f'nextContractDate': data[f'{contract_keys[n]}_nextContractDate']
                            }
                        }
                        contract_increment_data.update(new_data)
                elif key.startswith(f'new_ctr_'):
                    key = 0
                    if len(contract_keys) > 0:
                        key = int(contract_keys[-1][-2:])
                    if key < 9:
                        key = f'0{key + 1}'
                    else:
                        key = key + 1

                    new_data = {f'contract_{key}':
                        {
                            f'contractDate': data[f'new_ctr_contractDate'],
                            f'contractPeriod': data[f'new_ctr_contractPeriod'],
                            f'nextContractDate': data[f'new_ctr_nextContractDate']
                        }
                    }
                    contract_increment_data.update(new_data)

                else:
                    pass

        db.collection(companyname).document(u'employee').collection('employee').document(empid).update(contract_increment_data)
        if "increment_01" in contract_increment_data and len(contract_increment_data) >len(increment_data):
            inc_key = list(contract_increment_data.keys())
            print(inc_key[-1])

            contract_increment_data[inc_key[-1]]['empid'] = empid
            contract_increment_data[inc_key[-1]]['incrementDate']= (datetime.datetime.today().date()).strftime("%Y-%m-%d")

            doc_ref = db.collection(companyname).document('increments')
            data = contract_increment_data[inc_key[-1]]

            doc_ref.update({'increments': firestore.ArrayUnion([data])})
        elif "increment_01" in contract_increment_data :
            last_key = list(contract_increment_data.keys())[-1]
            last_record = contract_increment_data[last_key]
            effective_date = datetime.datetime.strptime(last_record['effectiveDate'], '%Y-%m-%d').date()
            today = datetime.datetime.today().date()
            print(today,effective_date)

            if effective_date <= today:
                doc_ref = db.collection(companyname).document('increments')
                increments = db.collection('alian_software').document('increments').get().to_dict()['increments']

                db.collection(companyname).document('employee').collection('employee').document(empid).update({'salary':round(float(last_record['total']))* 12})
                print("updated")
            else:
                doc_ref = db.collection(companyname).document('increments')
                increments = doc_ref.get().to_dict()['increments']
                inc_key = list(contract_increment_data.keys())
                contract_increment_data[inc_key[-1]]['updateIncrementDate'] = str(datetime.datetime.today().date())
                print(contract_increment_data)
                doc_ref = db.collection(companyname).document('increments')
                doc_ref.update({'increments': firestore.ArrayUnion([contract_increment_data[inc_key[-1]]])})
                print("update date")

    return redirect(url_for('employee_profile', id=empid, username=username))
@app.route('/<username>/tds_data_update/<id>', methods=['GET', 'POST'])
@login_required
def tds_data_update(username, id):
    """ UPDATE EMPLOYEE TDS DETAILS """
    if request.method == 'POST':
        form = request.get_json()
        update_obj.update_tds_info(companyname, form, id)
    return redirect(url_for('employee_profile', id=id, username=username))

@app.route('/delete_increment/<id>/<username>/<key>/<path:data_dict>/')
@login_required
def delete_increment(id, username,key, data_dict):
    # Convert the string representation of the dictionary back to a dictionary
    data = eval(data_dict)
    print(data,id,username,key)
    doc_ref=db.collection(companyname).document('increments')
    data.update({'empid':id})


    field_to_update = "increments"
    # Create a reference to the document
    doc_ref = db.collection(companyname).document("increments")
    doc = doc_ref.get()
    if doc.exists:
        # Get the current array from the document
        array_data = doc.to_dict().get('increments', [])
        # Remove the desired dictionary from the array
        modified_array = [item for item in array_data if item != data]
        print(modified_array)
        print(modified_array)
        # Update the document with the modified array
        doc_ref.set({'increments': modified_array}, merge=True)

        print("Dictionary deleted successfully from the array field.")
    else:
        print("Document does not exist.")


    db.collection(companyname).document('employee').collection('employee').document(id).update({key: firestore.DELETE_FIELD})

    return redirect(url_for('employee_profile',username=username,id=id))



@app.route('/<username>/department', methods=['GET', 'POST'])
@login_required
def department(username):
    """ DISPLAY DEPARTMENT """
    if request.method == 'POST':
        # Add New DEPARTMENT
        result = request.form
        dept.add_department(companyname, result)
    doc_ref = db.collection(str(companyname)).document(u'department')
    data = doc_ref.get().to_dict()
    return render_template('department.html', data=data, obj=dept, username=username)


@app.route('/<username>/delete_department/<dep> <pos>', methods=['GET', 'POST'])
@login_required
def delete_department(username, dep, pos):
    ''' DELETE DEPARTMENT '''
    a = dep, pos
    pattern = r'[^a-zA-Z\d\s]'
    # Use the re.sub() function to replace all occurrences of the pattern with an empty string
    dep = re.sub(pattern, '', dep)
    pos = re.sub(pattern, '', pos)
    dept.delete_department(companyname, dep, pos)
    return redirect(url_for('department', username=username))


@app.route("/<username>/set-storage-path/<salid>", methods=["POST"])
@login_required
def set_storage_path(username, salid):
    path = request.json["path"]
    salary = SalarySlip(db)
    salary.salary_slip(companyname, salid, path)
    return redirect(url_for('salary', username=username, salid=salid))



@app.route('/<username>/salary', methods=['GET', 'POST'])
@login_required
def salary(username):
    ''' DISPLAY SALARY DETAILS OF ALL MONTH IN YEAR '''
    holidays = db.collection(companyname).document('holidays').get().to_dict()
    if datetime.datetime.now().day == 15 and 'session_salary' not in session:
        SalaryCalculation(db, companyname).generate_salary(holidays=holidays)
        leaveobj.leave_add(companyname)
        session['session_salary'] = datetime.datetime.now()
        print("create session Variable")

    if request.method == 'POST':
        form = request.form
        data_dict = {}
        for key, value in form.items():
            if value != '':
                data_dict.update({key: value})
        # ADD Salary Criteria
        docs = db.collection(companyname).document('salary_calc')

        if len(docs.get().to_dict()) == 0:
            db.collection(companyname).document('salary_calc').set(data_dict)
        else:
            db.collection(companyname).document('salary_calc').update(data_dict)

    def get_salary_criteria():
        return db.collection(str(companyname)).document('salary_calc').get().to_dict()

    def get_all_month_salary_data():
        return Salarymanage(db).get_all_month_salary_data(companyname)

    def get_salary_status():
        return db.collection(companyname).document('salary_status').get().to_dict()

    with ThreadPoolExecutor(max_workers=3) as executor:
        salary_criteria_future = executor.submit(get_salary_criteria)
        salary_list_future = executor.submit(get_all_month_salary_data)
        salary_status_future = executor.submit(get_salary_status)

    salary_criteria = salary_criteria_future.result()
    salary_list = salary_list_future.result()
    salary_status = salary_status_future.result()
    year = datetime.datetime.now().year
    return render_template('salary_sheet_month.html', data=salary_list, salary_criteria=salary_criteria
                           , username=username, salary_status=salary_status, year=year)



@app.route('/<username>/upload/<salid>', methods=['POST'])
@login_required
def upload(username,salid):
    ''' IMPORT EXCEL SHEET FOR SALARY DATA '''
    print(salid)
    if request.method == 'POST':
        file = request.files['file']
        all_data = read_excel_leave_data.read_excel_rows(file)
        # GET ALL EMPLOYEE USER ID FROM EXCEL SHEET
        for data in all_data:
            # print(data)
            new_data = db.collection(companyname).document(u'employee').collection('employee').where('cosecID', '==', data[" User ID"]).get()
            if len(new_data) != 0:
                for details in new_data:
                    document_name = details.to_dict()['userID']
                    emp_salary_data = {'cosecID': data[" User ID"], 'WO': data["WO"], 'UL': data["UL"],
                                       'OT': data["OT"], 'WrkHrs': data[" WrkHrs"],'paidLeave':data["PL"], 'CL': data["C_L"],
                                       'PL': data["P_L"], 'SL': data["S_L"]}
                    print(emp_salary_data)
                    try:
                        db.collection(companyname).document(u'employee').collection('employee').document(document_name).collection('salaryslips').document(salid).update(emp_salary_data)
                    except:
                        pass
                    # print(details.to_dict())
        return redirect(url_for('salary', username=username))
    return redirect(url_for('salary', username=username))



@app.route('/<username>/salarysheetview/<salid>', methods=['GET', 'POST'])
@login_required
def salary_sheet_view(username, salid):
    # month = int(salid[5:])

    if request.method == 'POST':
        form = request.form
        print(form)
        if 'date' in form.keys():
            form = form.to_dict()
            data = {str(form["date"]): form["description"]}
            print(data)
            db.collection(companyname).document('holidays').update(data)
            holidays = db.collection(companyname).document('holidays').get().to_dict()
            moath_data = moth_count.count(holidays)
            db.collection(companyname).document('month_data').set(moath_data)
        elif 'date' not in form.keys():
            fields = {}
            for key, value in form.items():
                fields.update({key: value})

            path = get_download_folder()
            salary_excel = SalaryData(db)
            salary_excel.add_data(salid=salid, fields=fields, path=path, companyname=companyname)
        else:
            pass

    ''' DISPLAY SALARY DETAILS OF EMPLOYEES IN MONTH '''
    holidays = db.collection(companyname).document('holidays').get().to_dict()
    moath_data = moth_count.count_previous_month(holidays=holidays, salid=salid)
    working_days = moath_data['workingDays']

    holidays = moth_count.get_holidays(holidays)

    salary_list = Salarymanage(db).get_all_emp_salary_data(companyname, salid)

    salary_status = db.collection(companyname).document('salary_status').get()
    salary_status = salary_status.get(datetime.date(1900, int(salid[4:]), 1).strftime('%B'))

    return render_template('salary_sheet_view.html', data=salary_list, salid=salid, username=username,
                           salary_status=salary_status, moath_data=moath_data, holidays=holidays)

@app.route('/<username>/salarysheetedit/<empid> <salid>', methods=['GET', 'POST'])
@login_required
def salary_sheet_edit_(username, empid, salid):
    ''' EDIT SALARY DETAILS OF EMPLOYEE IN MONTH '''
    if request.method == 'POST':
        result = request.form
        Salarymanage(db).salary_update(companyname, empid, salid, data=result)
        return redirect(url_for('salary_sheet_view', salid=salid, username=username))

    holidays = db.collection(companyname).document('holidays').get().to_dict()
    moath_data = moth_count.count_previous_month(holidays=holidays, salid=salid)
    working_days = moath_data['workingDays']
    employee_salary_data = Salarymanage(db).get_salary_data(companyname, empid, salid)
    salary_percentage = (db.collection(companyname).document('salary_calc').get()).to_dict()
    return render_template('salary_sheet_edit_personal.html', data=employee_salary_data, id=salid
                           , salary_data=salary_percentage, username=username,
                           working_days=working_days)


@app.route('/<username>/salarysheeteditall/<salid>', methods=["GET","POST"])
@login_required
def salary_sheet_edit_all(username, salid):
    holidays = db.collection(companyname).document('holidays').get().to_dict()
    moath_data = moth_count.count_previous_month(holidays=holidays, salid=salid)
    working_days = moath_data['workingDays']
    salary_percentage = (db.collection(companyname).document('salary_calc').get()).to_dict()
    salary_list = Salarymanage(db).get_all_emp_salary_data(companyname, salid)
    # print(salary_list)
    return render_template("salary_sheet_edit_all.html", datas=salary_list, username=username, salid=salid,
                           working_days=working_days, salary_data=salary_percentage)


@app.route('/save_edited_data', methods=['POST'])
@login_required
def save_edited_data():
    # Get the form data from request.form
    form_data = request.form
    # form_dict = form_data.to_dict()
    form_dict = dict(form_data)
    # print(f'data  = {form_dict}')
    salid = request.args.get('sal_id')
    username = request.args.get('user_name')
    # print(f"{salid} is salid and {username} is username")

    print(form_dict)

    all_data = {}

    for key, value in form_dict.items():
        if value == '':
            value = 0
        document_name = key.split('_')[0]  # Extracting the document name from the key
        field_name = key.split('_')[1]  # Extracting the field name from the key
        # If key not exist then add new key
        if len(all_data) == 0 or document_name not in all_data.keys():
            all_data.update({document_name: {field_name: value}})
        # If key already exist then add field in key value
        else:
            all_data[document_name].update({field_name: value})

    # Updating the document in Firestore
    for emp_id, sal_data in all_data.items():
        db.collection(companyname).document('employee').collection('employee').document(emp_id).collection('salaryslips').document(salid).update(sal_data)

    # Return a response to the client
    return redirect(url_for('salary_sheet_view', username=username, salid=salid))


@app.route('/<username>/set_status/<salid>/<status>')
@login_required
def set_status(username, salid, status):
    ''' SALARY SLIP PDF GENERATION '''
    month = datetime.date(1900, int(salid[3:]), 1).strftime('%B')
    status = status
    data = {month: status}
    salary_status = db.collection(companyname).document('salary_status').update(data)
    return redirect(url_for('salary_sheet_view', username=username, salid=salid))




def get_download_folder():
    if os.name == 'nt':  # for Windows
        import winreg
        sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
        downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
            location = winreg.QueryValueEx(key, downloads_guid)[0]
    elif os.name == 'darwin':  # for macOS
        location = os.path.expanduser('~/Downloads')
    else:  # for Linux/Unix
        location = os.path.expanduser('~/Downloads')
    return location

from flask import Flask, make_response


@app.route('/download_pdf')
@login_required
def download_pdf():
    # Open the PDF file
    with open('sample.pdf', 'rb') as file:
        pdf_content = file.read()

    # Create a response object with the PDF content
    response = make_response(pdf_content)
    response.headers['Content-Disposition'] = 'attachment; filename=example.pdf'
    response.headers['Content-Type'] = 'application/pdf'

    return response




@app.route('/<username>/pdf/<salid>')
@login_required
def pdf_all(username, salid):
    ''' SALARY SLIP PDF GENERATION '''
    salary_list = Salarymanage(db).get_all_emp_salary_data(salid=salid, companyname=companyname)
    print(salary_list)
    responses = []
    for i in salary_list:
        empid = salary_list[i]["userID"]
        path = get_download_folder()
        salary = SalarySlip(db)
        responcedata = salary.salary_slip_personal(companyname, empid, salid, path)
        return responcedata
    # for response in responses:
    #     # CHECK THE USER
    #     return response
    return redirect(url_for('salary', username=username, salid=salid))



# @app.route('/<username>/excel/<salid>')
# def excel_for_bank(companyname,username, salid):
#     ''' GENERATE EXCELSHEET FOR BANK '''
#     if 'storage_path' in session:
#         path = session["storage_path"]
#         salary_excel = SalaryData(db)
#         salary_excel.add_data(companyname, salid, path)
#     return redirect(url_for('salary_sheet_view', salid=salid, username=username))


# @app.route('/tds/<id>', methods=['GET', 'POST'])
# def tds(companyname, id):
#     ''' DISPLAY TDS DETAILS OF EMPLOYEE '''
#     profile = Profile(companyname, db, id)
#     employee_tds_data = {'personal_data': profile.personal_data(), 'tds_data': profile.tds_data()}
#     return render_template('tds_test.html', data=employee_tds_data, )

@app.route('/<username>/add_data', methods=['POST', 'GET'])
@login_required
def add_data(username):
    if request.method == 'POST':
        data = request.form

        data = data.to_dict()

        excel_path = data['path']

        excel = ExcelData(db)
        excel.store_excel_data(companyname, excel_path)
        return redirect(url_for('dashboard', username=username))
    return render_template('add_excel_file.html', username=username)


@app.route('/<username>/send_email/<salid>')
@login_required
def send_employee_salaryslip(username, salid):
    ''' GENERATE EXCELSHEET FOR BANK '''

    path = get_download_folder()
    auth_data = db.collection(companyname).document('admin').get().to_dict()
    company_mail = auth_data['AdminID']
    auth_password = auth_data['auth_password']
    docs = db.collection(companyname).document(u'employee').collection('employee').stream()
    employee_list = {}
    for doc in docs:
        employee_list.update({doc.id: doc.to_dict()})
    for key, value in employee_list.items():
        data = value

        mail_obj.send_employee_pdf(company_mail=company_mail, data=data, auth_password=auth_password, path=path,
                                   companyname=companyname)
    return redirect(url_for('salary_sheet_view', salid=salid, username=username))


if __name__ == '__main__':
    # app.run(debug=True, port=300)
    app.run(debug=True, host="0.0.0.0", port=3005)
