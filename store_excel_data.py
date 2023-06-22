import json

import requests
from openpyxl import load_workbook
from datetime import datetime
from firebase_admin import auth

FIREBASE_WEB_API_KEY = "AIzaSyDe2qwkIds8JwMdLBbY3Uw7JQkFRNXtFqo"
rest_api_url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword"



def sign_in_with_email_and_password(email: str, password: str, return_secure_token: bool = True):
    payload = json.dumps({
        "email": email,
        "password": password,
        "returnSecureToken": return_secure_token
    })

    r = requests.post(rest_api_url,
                      params={"key": FIREBASE_WEB_API_KEY},
                      data=payload)

    return r.json()


class Uploaddata():

    def __init__(self, db):
        self.db = db

    def read_excel_rows(self, file_path, companyname):
        workbook = load_workbook(filename=file_path)

        sheet = workbook['Sheet1']

        max_row = sheet.max_row

        data = []

        for row in range(1, max_row + 1):
            row_data = []
            for column in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=column).value
                # Check if the cell value is a datetime object

                if isinstance(cell_value, datetime):
                    # Convert the datetime to the desired format
                    cell_value = cell_value.strftime("%d-%m-%Y")

                if cell_value != None:
                    row_data.append(cell_value)

            data.append(row_data)

        header_row_index = None

        for i, row in enumerate(data):
            if row == ['Name', 'Department', 'Job Position', 'Designation', 'COSEC ID', 'Password', 'Confirm Password', 'DOJ (DD-MM-YYYY)', 'DOB (DD-MM-YYYY)', 'CTC ₹ per year', 'Email', 'role', 'Bank Name ', 'Account Nuimber', 'Location', 'EPFO', 'ABRY']:
                header_row_index = i

        if header_row_index is None:
            #print("Invalid data format. Unable to find header row or attendance data.")
            exit()
        #
        column_headers = data[header_row_index]

        employee_records = []

        for row in data:
            if row != []:
                record = dict(zip(column_headers, row))
                employee_records.append(record)

        all_data = employee_records
        count = 0

        for data in all_data[1:]:
            doj = ''
            if data['DOJ (DD-MM-YYYY)'][2] == "/":
                date = (data['DOJ (DD-MM-YYYY)']).split('/')
                doj = f"{date[2]}-{date[1]}-{date[0]}"
            elif data['DOJ (DD-MM-YYYY)'][2] == "-":
                date = (data['DOJ (DD-MM-YYYY)']).split('-')
                doj = f"{date[2]}-{date[1]}-{date[0]}"



            dob = ''
            if data['DOB (DD-MM-YYYY)'][2] == "/":
                date = (data['DOB (DD-MM-YYYY)']).split('/')
                dob = f"{date[2]}-{date[1]}-{date[0]}"
            elif data['DOB (DD-MM-YYYY)'][2] == "-":
                date = (data['DOB (DD-MM-YYYY)']).split('-')

                dob = f"{date[2]}-{date[1]}-{date[0]}"
            emp_data = {'cosecID': data["COSEC ID"], 'employeeName': data["Name"], 'department': data["Department"],
                        'designation': data["Designation"],'jobPosition': data['Job Position'], "password": data['Password'],
                        'confirmPassword': data["Confirm Password"], 'doj': doj,'dob':dob,
                        'salary': data['CTC ₹ per year'], 'workEmail': data['Email'], 'role':   data['role'],
                        'bankName': data['Bank Name '],'accountNumber': data['Account Nuimber'],
                        'location': data['Location'], 'epfo_status': data['EPFO'], 'abry': data['ABRY']
                        }
            # #print(emp_data)
            count+=1


            # try:
            #     user = auth.create_user(
            #         email=emp_data['workEmail'],
            #         password=emp_data['password']
            #     )
            #     emp_data["userID"] = user.uid
            #     print("in try")
            #     self.db.collection(companyname).document('employee').collection('employee').document(user.uid).set(emp_data)
            # except Exception as e:
            #     user_auth = sign_in_with_email_and_password(email=emp_data['workEmail'].strip(),
            #                                                 password=emp_data['password'])
            #     if "registered" in user_auth:
            #         emp_data["userID"] = user_auth['localId']
            #
            #         self.db.collection(companyname).document('employee').collection('employee').document(
            #             user_auth['localId']).set(emp_data)
            #         print("in excep")
            #     else:
            #         # global counts
            #         # counts += 1
            #         print("in else")
            #         emp_data["userID"] = data["Name"].strip()
            #         self.db.collection(companyname).document('employee').collection('employee').document(emp_data['userID']).set(emp_data)

            self.db.collection(companyname).document('employee').collection('employee').document().set(emp_data)

